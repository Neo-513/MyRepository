import openpyxl
import re
import xlrd
import xlwt


def read(path):
	"""读取excel
	:param path: 文件路径
	:return: 键名为页名、键值为二维数组的数据集合字典
	"""
	filetype = path.split(".")[-1].lower()
	if filetype == "xlsx":
		return _read_xlsx(path)
	elif filetype == "xls":
		return _read_xls(path)


def write(path, dic):
	"""读取excel
	:param path: 文件路径
	:param dic: 键名为页名、键值为二维数组的数据集合字典
	"""
	filetype = path.split(".")[-1].lower()
	if filetype == "xlsx":
		_write_xlsx(path, dic)
	elif filetype == "xls":
		_write_xls(path, dic)


def _read_xlsx(path):
	file = openpyxl.load_workbook(path)  # 文件对象
	dic = {}  # 数据集合字典

	for sheetname in file.sheetnames:  # 遍历表名
		datas = file[sheetname]  # 单张表数据集合
		dic[sheetname] = [[cell.value if cell.value else "" for cell in row] for row in datas.rows]

	file.close()  # 关闭文件对象
	print(f"[READ]    {path}")  # 打印信息
	return dic


def _write_xlsx(path, dic):
	file = openpyxl.Workbook()  # 文件对象
	file.remove(file.active)  # 删除默认表

	for sheetname in dic:  # 遍历表名
		sheet = file.create_sheet(sheetname)  # 创建表
		datas = dic[sheetname]  # 单张表数据集合
		row = len(datas)  # 行数
		col = 0 if row == 0 else len(datas[0])  # 列数
		for r in range(row):  # 遍历行
			for c in range(col):  # 遍历单元格
				sheet.cell(r + 1, c + 1).value = datas[r][c]  # 写入单元格数据（起始索引为1）

	file.save(path)  # 保存文件
	file.close()  # 关闭文件对象
	print(f"[WITRE]    {path}")  # 打印信息


def _read_xls(path):
	file = xlrd.open_workbook(path)  # 文件对象
	dic = {}  # 数据集合字典

	for sheetname in file.sheet_names():  # 遍历表名
		datas = file.sheet_by_name(sheetname)  # 单张表数据集合
		dic[sheetname] = [[
			int(cell.value) if re.findall("^\d+\.0*$", str(cell.value)) else cell.value for cell in row
		] for row in datas.get_rows()]

	print(f"[READ]    {path}")  # 打印信息
	return dic


def _write_xls(path, dic):
	file = xlwt.Workbook()  # 文件对象

	for sheetname in dic:  # 遍历表名
		sheet = file.add_sheet(sheetname)  # 创建表
		datas = dic[sheetname]  # 单张表数据集合
		row = len(datas)  # 行数
		col = 0 if row == 0 else len(datas[0])  # 列数
		for r in range(row):  # 遍历行
			for c in range(col):  # 遍历单元格
				sheet.write(r, c, datas[r][c])  # 写入单元格数据（起始索引为0）

	file.save(path)  # 保存文件
	print(f"[WITRE]    {path}")  # 打印信息
